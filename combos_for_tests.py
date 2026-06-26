"""
SPREADSHEET COMBINER
=====================
PURPOSE:
    Reads all CSV files from a single folder, adds a "pocketlab_number" column
    to each (pulled from the filename, e.g. "0618_TEST2_PL1.csv" → 1), and
    combines everything into one Excel file saved in the same folder.

    Files with "KESTREL" anywhere in their name are automatically skipped.

HOW TO USE THIS SCRIPT:
    1. Fill in Section 1 below (the only section you need to edit)
    2. Run the script:   python combine_spreadsheets.py
    3. The combined output file will be saved inside the same folder as your CSVs,
       automatically named:  COMBINED_<folder name>.xlsx

DEPENDENCIES — install once if not already installed:
    pip install pandas openpyxl
"""

import os
import re
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# =============================================================================
# SECTION 1: CONFIGURATION — THIS IS THE ONLY SECTION YOU NEED TO EDIT
# =============================================================================

INPUT_FOLDER = r"C:\Users\dresd\Downloads\0618_R1000"
                            # Path to the folder containing the CSV files to combine
                            # All .csv files in this folder will be read and merged
                            # (except any with "KESTREL" in the filename)

SHEET_NAME = "Combined"
                            # Name of the sheet in the output Excel file
                            # Keep it short — Excel allows a maximum of 31 characters


# =============================================================================
# END OF USER INPUT — DO NOT EDIT BELOW THIS LINE UNLESS CUSTOMIZING
# =============================================================================


# =============================================================================
# SECTION 2: VALIDATE INPUTS AND BUILD OUTPUT FILE PATH
# =============================================================================

if not os.path.isdir(INPUT_FOLDER):
    print(f"ERROR: Input folder not found: {INPUT_FOLDER}")
    sys.exit(1)

# Output file is saved inside the same folder, named after the folder itself
# Example: folder "0618_R1000" → output file "COMBINED_0618_R1000.xlsx"
folder_name = os.path.basename(os.path.normpath(INPUT_FOLDER))
OUTPUT_FILE = os.path.join(INPUT_FOLDER, f"COMBINED_{folder_name}.xlsx")


# =============================================================================
# SECTION 3: FIND ALL CSV FILES, SKIPPING KESTREL FILES
# =============================================================================

all_files_in_folder = sorted([
    f for f in os.listdir(INPUT_FOLDER)
    if f.lower().endswith(".csv")
])

csv_files = []
skipped   = []

for f in all_files_in_folder:
    if re.search(r"KESTREL", f, re.IGNORECASE):  # BUG FIX: was r"PL" — caused PocketLab files to be skipped instead of Kestrel files
        skipped.append(f)
    else:
        csv_files.append(os.path.join(INPUT_FOLDER, f))

if skipped:
    print(f"=== Skipping {len(skipped)} KESTREL file(s) ===")
    for f in skipped:
        print(f"  SKIPPED: {f}")
    print()

if not csv_files:
    print(f"ERROR: No usable CSV files found in folder: {INPUT_FOLDER}")
    sys.exit(1)

print(f"=== Found {len(csv_files)} CSV file(s) to combine ===")
for f in csv_files:
    print(f"  {os.path.basename(f)}")
print()


# =============================================================================
# SECTION 4: EXTRACT POCKETLAB NUMBER FROM FILENAME
# =============================================================================
# The PocketLab number is the digits that follow "_PL" in the filename.
# Example: "0618_TEST2_PL1.csv" → pocketlab_number = 1
#          "0618_TEST2_PL14.csv" → pocketlab_number = 14

def extract_pocketlab_number(filepath):
    """
    Returns the PocketLab number (integer) from a CSV filename.
    Returns None if no _PL followed by digits is found.
    """
    filename = os.path.basename(filepath)
    match = re.search(r"_PL(\d+)", filename, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return None


# =============================================================================
# SECTION 5: READ ALL FILES AND ADD POCKETLAB NUMBER COLUMN
# =============================================================================
# Column order and extra columns (e.g. Latitude/Longitude) may differ between
# devices — that's fine. We align everything by column NAME at combine time,
# not by column position. Missing columns get filled with blanks automatically.

print("Reading files...")

frames = []

for filepath in csv_files:
    filename = os.path.basename(filepath)

    # --- Extract PocketLab number ---
    pl_number = extract_pocketlab_number(filepath)
    if pl_number is None:
        print(f"  WARNING: Could not find a PocketLab number (_PL#) in '{filename}' — skipping.")
        continue

    # --- Read the CSV ---
    try:
        df = pd.read_csv(filepath, dtype=str)
    except Exception as e:
        print(f"  ERROR reading '{filename}': {e}")
        print("  Skipping this file.")
        continue

    if df.empty:
        print(f"  WARNING: '{filename}' is empty — skipping.")
        continue

    # --- Add pocketlab_number as the first column ---
    df.insert(0, "pocketlab_number", pl_number)

    print(f"  {filename}: PocketLab {pl_number}, {len(df)} row(s), cols: {list(df.columns)}")
    frames.append(df)

print()

if not frames:
    print("ERROR: No files could be read successfully. Nothing to combine.")
    sys.exit(1)


# =============================================================================
# SECTION 6: COMBINE ALL DATA INTO ONE TABLE
# =============================================================================

combined = pd.concat(frames, ignore_index=True)

print(f"=== Combined total: {len(combined)} rows across {len(frames)} file(s) ===")
print()


# =============================================================================
# SECTION 7: SAVE TO EXCEL WITH FORMATTING
# =============================================================================

print(f"Saving combined file to: {OUTPUT_FILE}")

combined.to_excel(OUTPUT_FILE, sheet_name=SHEET_NAME, index=False)

# --- Apply formatting with openpyxl ---
wb = load_workbook(OUTPUT_FILE)
ws = wb[SHEET_NAME]

header_font   = Font(name="Arial", bold=True)
header_fill   = PatternFill("solid", fgColor="BDD7EE")   # Light blue
header_align  = Alignment(horizontal="center")
header_border = Border(bottom=Side(style="thin"))

for cell in ws[1]:   # Row 1 = headers
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = header_align
    cell.border    = header_border

# Apply Arial font to all data rows
data_font = Font(name="Arial")
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = data_font

ws.freeze_panes = "A2"   # Freeze header row so it stays visible when scrolling

# Auto-size columns to fit content
for col in ws.columns:
    max_len = max(
        len(str(cell.value)) if cell.value is not None else 0
        for cell in col
    )
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

wb.save(OUTPUT_FILE)

print()
print(f"=== Done! Combined file saved to: {OUTPUT_FILE} ===")
print(f"    Sheet : {SHEET_NAME}")
print(f"    Rows  : {len(combined)}")
print(f"    Cols  : {', '.join(combined.columns)}")