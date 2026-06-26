"""
HEAT MAPPING DATA COMBINING SCRIPT
==================================
PURPOSE:
    Reads all CSV session files from a single folder, automatically detects:
      - The date and route number (from the folder name, e.g. "0608_R30")
      - Each collector's initials (from each CSV filename, e.g. "0608_R30_BH_T1.csv" → "BH")
      - Whether each file is morning or afternoon data (from the Time column)

    Produces one Excel output file per collector per time block:
      e.g.  combined_R30_06082026_BH_AM_noGPS.xlsx
            combined_R30_06082026_BH_PM_noGPS.xlsx
            combined_R30_06082026_TJ_AM_noGPS.xlsx
            combined_R30_06082026_TJ_PM_noGPS.xlsx

HOW TO USE THIS SCRIPT:
    1. Fill in Section 1 below (the only section you need to edit)
    2. Run the script:   python combining_raw_including_missing_gps.py
    3. Output Excel files will be saved to the folder you specify in OUTPUT_DIR

DEPENDENCIES — install once if not already installed:
    pip install pandas openpyxl
"""

import os
import re
import sys
import pandas as pd
from datetime import datetime, time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# =============================================================================
# SECTION 1: CONFIGURATION — THIS IS THE ONLY SECTION YOU NEED TO EDIT
# =============================================================================

# --- Route Information -------------------------------------------------------

ROUTE_NAME    = "williams_delight_e"
                            # Full route name as used in data org system
                            # (e.g., "st_thomas_town", "hull_bay", "uvi_lindbergh")

ROUTE_ABBREV  = "williams_e"
                            # Short abbreviation used in route_id
                            # (e.g., "stt", "hull", "uvi")

ROUTE_NUMBER  = 17
                            # Route number as an integer (e.g., 6, 30)
                            # This should match the number in your folder name

WALK_INSTANCE = "a"
                            # "a" = first time this route was walked
                            # "b" = second time this route was walked
                            # (must be a different day ~ only change this
                            # if the route had to be re-walked another day)

YEAR          = "2026"
                            # The year data was collected, as a 4-digit string
                            # (e.g., "2025", "2026")

# --- Folder Paths ------------------------------------------------------------

SESSION_FOLDER = r"C:\Users\dresd\USVI_2026\raw_data\0607_R17"
                            # Path to the folder containing all your session CSV files
                            # The folder name MUST follow the format: {MMDD}_R{route number}
                            # Example: "C:\Users\dresd\USVI_2026\raw_data\0615_R06"

OUTPUT_DIR_1     = r"C:\Users\dresd\USVI_2026\combined_data\0607_R17\all_R17"
                            # Folder where intermediate Excel output files will be saved
                            # So split up by person by time block
                            # The folder must already exist before running this script

OUTPUT_DIR = r"C:\Users\dresd\USVI_2026\combined_data\0607_R17"
                            # Folder where the combined Excel file will be saved
                            # This folder must already exist before running the script


# =============================================================================
# END OF USER INPUT — DO NOT EDIT BELOW THIS LINE UNLESS CUSTOMIZING
# =============================================================================


# =============================================================================
# SECTION 2: PARSE DATE AND ROUTE INFO FROM FOLDER NAME
# =============================================================================
# The folder name is expected to follow the format: {MMDD}_R{route number}
# Example: "0608_R30" → month=06, day=08, route=30

folder_name = os.path.basename(os.path.normpath(SESSION_FOLDER))

folder_match = re.match(r"^(\d{2})(\d{2})_R(\d+)$", folder_name)
if not folder_match:
    print(f"ERROR: Folder name '{folder_name}' does not match expected format MMDD_R{{route number}}")
    print("       Example of a valid folder name: 0608_R30")
    sys.exit(1)

month_str = folder_match.group(1)   # e.g., "06"
day_str   = folder_match.group(2)   # e.g., "08"
DATE      = f"{int(month_str)}/{day_str}/{YEAR}"   # e.g., "6/08/2026"

# Build the route_id used as a label in the output data
ROUTE_ID = f"{ROUTE_ABBREV}_{WALK_INSTANCE}_{ROUTE_NUMBER}"

# Build a date string for use in filenames: MMDDYYYY
date_str_for_filename = f"{month_str}{day_str}{YEAR}"

print("=== Parsed from folder name ===")
print(f"  Folder name   : {folder_name}")
print(f"  Date          : {DATE}")
print(f"  Route ID      : {ROUTE_ID}")
print(f"  Route Number  : {ROUTE_NUMBER}")
print()


# =============================================================================
# SECTION 3: FIND ALL CSV FILES IN THE SESSION FOLDER
# =============================================================================
# Files matching any of the patterns below will be automatically skipped:
#   - Filenames containing "KESTREL" or "Kestrel" (any capitalization)
#   - Filenames containing "WBGT" (any capitalization)
#   - Filenames containing "M" followed immediately by one or more digits
#     anywhere in the name (e.g., M1, M12, M3)

def should_skip_file(filename):
    """
    Returns True if this CSV file should be ignored.
    Checks for Kestrel, WBGT, and M+number patterns in the filename.
    """
    name_no_ext = os.path.splitext(filename)[0]   # Remove .csv before checking

    # Skip any file with "kestrel" or "wbgt" in the name (case-insensitive)
    if re.search(r"kestrel|wbgt", name_no_ext, re.IGNORECASE):
        return True

    # Skip files where any underscore-separated part is exactly M followed by digits
    # (e.g., a part like "M1", "M12" — but NOT a part like "MF" or "M" alone)
    parts = name_no_ext.split("_")
    for part in parts:
        if re.fullmatch(r"M\d+", part, re.IGNORECASE):
            return True

    return False


if not os.path.isdir(SESSION_FOLDER):
    print(f"ERROR: Session folder not found: {SESSION_FOLDER}")
    sys.exit(1)

all_files_in_folder = sorted([
    f for f in os.listdir(SESSION_FOLDER)
    if f.lower().endswith(".csv")
])

# Separate files into those we'll use and those we'll skip
all_csvs  = []
skipped   = []

for f in all_files_in_folder:
    if should_skip_file(f):
        skipped.append(f)
    else:
        all_csvs.append(os.path.join(SESSION_FOLDER, f))

# Report what was skipped
if skipped:
    print(f"=== Skipping {len(skipped)} file(s) (Kestrel / WBGT / M+number) ===")
    for f in skipped:
        print(f"  SKIPPED: {f}")
    print()

if not all_csvs:
    print(f"ERROR: No usable CSV files found in folder: {SESSION_FOLDER}")
    print("       (All files may have been skipped, or the folder may be empty.)")
    sys.exit(1)

print(f"=== Found {len(all_csvs)} usable CSV file(s) in folder ===")
for f in all_csvs:
    print(f"  {os.path.basename(f)}")
print()


# =============================================================================
# SECTION 4: EXTRACT INITIALS FROM EACH CSV FILENAME
# =============================================================================
# Three methods are tried in order:
#
# METHOD 0 — PocketLab number (checked first, before any initials logic):
#   If any underscore-separated part of the filename matches "PL" followed by
#   one or more digits (case-insensitive), that whole part is returned as-is.
#   "PL" is never used as personal initials, so any occurrence of PL+number
#   is unambiguously a PocketLab device identifier.
#   Example: "0621_R27_PL1_T1.csv"  → "PL1"
#            "0621_R27_PL12_T2.csv" → "PL12"
#
# METHOD 1 — Standard format: {MMDD}_R{route}_{INITIALS}_T{number}.csv
#   Initials are the part between the 2nd and 3rd underscore.
#   Example: "0608_R30_BH_T1.csv" → "BH"
#   This method is used when the filename has 3 or more underscore-separated parts
#   AND the third part consists entirely of letters (no digits mixed in).
#
# METHOD 2 — Fallback for non-standard filenames:
#   Scan the filename (without extension) for the first occurrence of exactly
#   two consecutive uppercase letters not surrounded by other letters.
#   Example: "0607_R20_MF3.csv" → "MF"
#   This handles files where the initials are embedded in a part like "MF3".

def extract_initials(filepath):
    """
    Try to extract 2-letter collector initials from a CSV filename.
    Returns the initials as an uppercase string, or None if none can be found.
    """
    filename    = os.path.basename(filepath)
    name_no_ext = os.path.splitext(filename)[0]   # Remove .csv
    parts       = name_no_ext.split("_")

    # --- Method 0: PocketLab number ---
    # Scan all parts for "PL" followed by one or more digits.
    # Checked before initials logic because "PL" is never a valid set of initials.
    for part in parts:
        if re.fullmatch(r"PL\d+", part, re.IGNORECASE):
            return part.upper()  # e.g. "pl1" → "PL1"

    # --- Method 1: Standard format ---
    # The third part (index 2) should be pure letters, e.g. "BH" or "TJ"
    if len(parts) >= 3 and re.fullmatch(r"[A-Za-z]+", parts[2]):
        return parts[2].upper()

    # --- Method 2: Fallback — find two consecutive uppercase letters in filename ---
    # Search each underscore-separated part for a run of exactly 2 letters
    # that appear at the start of the part or after any digits.
    # Example: "MF3" → finds "MF"; "T14" → no match (only one letter before digits)
    for part in parts:
        match = re.search(r"(?<![A-Za-z])([A-Za-z]{2})(?![A-Za-z])", part)
        if match:
            return match.group(1).upper()

    # Could not find initials by either method
    return None


csv_by_initials = {}   # { "BH": [file1, file2, ...], "TJ": [...], ... }

for filepath in all_csvs:
    initials = extract_initials(filepath)
    if initials is None:
        print(f"WARNING: Could not extract initials from filename '{os.path.basename(filepath)}' — skipping.")
        continue
    csv_by_initials.setdefault(initials, []).append(filepath)

print("=== Files grouped by collector initials ===")
for initials, files in csv_by_initials.items():
    print(f"  {initials}: {len(files)} file(s)")
    for f in files:
        print(f"    {os.path.basename(f)}")
print()


# =============================================================================
# SECTION 5: DEFINE HELPER — CLASSIFY A CSV AS MORNING OR AFTERNOON
# =============================================================================
# Rule:
#   If the Time column values contain "AM", OR the time is before 1:30 PM → morning
#   If the Time column values contain "PM" AND the time is at or after 1:30 PM → afternoon
#
# Since each CSV is assumed to contain only one time block, we check the first
# non-empty Time value in the file to determine its classification.

CUTOFF_TIME = time(13, 30, 0)   # 1:30:00 PM — times at or after this are PM

def classify_time_block(filepath):
    """
    Read the Time column from a CSV and return "morning" or "afternoon".
    Returns None if the time cannot be determined.
    """
    try:
        df = pd.read_csv(filepath, dtype=str, usecols=["Time"])
    except Exception as e:
        print(f"WARNING: Could not read Time column from '{os.path.basename(filepath)}': {e}")
        return None

    # Find first non-empty Time value
    for raw_val in df["Time"].dropna():
        raw_val = str(raw_val).strip()
        if not raw_val:
            continue

        # Check for AM or PM in the string
        upper_val = raw_val.upper()
        if "AM" in upper_val:
            return "morning"
        elif "PM" in upper_val:
            # Parse the time and compare to 1:30 PM cutoff
            try:
                parsed = datetime.strptime(raw_val, "%I:%M:%S %p").time()
            except ValueError:
                try:
                    parsed = datetime.strptime(raw_val, "%I:%M %p").time()
                except ValueError:
                    return "afternoon"   # Has "PM" but can't parse — assume afternoon
            if parsed >= CUTOFF_TIME:
                return "afternoon"
            else:
                return "morning"
        else:
            # No AM/PM — assume 24-hour format; compare numerically
            try:
                parsed = datetime.strptime(raw_val, "%H:%M:%S").time()
            except ValueError:
                try:
                    parsed = datetime.strptime(raw_val, "%H:%M").time()
                except ValueError:
                    continue   # Can't parse; try next row
            if parsed >= CUTOFF_TIME:
                return "afternoon"
            else:
                return "morning"

    print(f"WARNING: Could not determine time block for '{os.path.basename(filepath)}'")
    return None


# =============================================================================
# SECTION 6: DEFINE HELPER — PARSE TIME VALUE TO 24-HOUR FORMAT
# =============================================================================

def parse_time_to_24h(t):
    t = str(t).strip()
    if re.search(r"AM|PM", t, re.IGNORECASE):
        try:
            return pd.to_datetime(t, format="%I:%M:%S %p").strftime("%H:%M:%S")
        except ValueError:
            try:
                return pd.to_datetime(t, format="%I:%M %p").strftime("%H:%M:%S")
            except ValueError:
                return t
    return t   # Already in HH:MM:SS format


# =============================================================================
# SECTION 7: DEFINE EXPECTED RAW COLUMN NAMES
# =============================================================================

RAW_COLUMNS = {
    "Date":                                        "raw_date",
    "Time":                                        "raw_time",
    "Elapsed Seconds":                             "raw_elapsed",
    "Latitude":                                    "raw_lat",
    "Longitude":                                   "raw_lon",
    "Temperature Probe-Temperature (&#x2103;)":    "raw_temp",
    "Humidity-Humidity (%RH)":                     "raw_humidity",
    "Heat Index-Heat Index (&#x2103;)":            "raw_heat_index",
    "Dew Point-Dew Point (&#x2103;)":              "raw_dew_point",
    "Ambient Light Intensity-Light (lux)":         "raw_light",
}


# =============================================================================
# SECTION 8: DEFINE HELPER — READ, CLEAN, AND BUILD OUTPUT FOR A LIST OF FILES
# =============================================================================

def process_files(file_list, initials, time_block):
    """
    Reads a list of CSV files, cleans the data, and returns a formatted DataFrame.

    Parameters:
        file_list  : list of file paths to read
        initials   : collector initials string (e.g., "BH")
        time_block : "morning" or "afternoon"

    Returns:
        A cleaned pandas DataFrame ready to write to Excel, or None on failure.
    """

    time_label   = "AM"     if time_block == "morning" else "PM"

    print(f"  Processing {len(file_list)} file(s) for {initials} ({time_label})...")

    # --- Read and combine all CSVs ---
    frames = []
    for filepath in file_list:
        df = pd.read_csv(filepath, dtype=str)

        # Add missing GPS columns as blank (some loggers don't record GPS)
        if "Latitude" not in df.columns:
            df["Latitude"] = None
        if "Longitude" not in df.columns:
            df["Longitude"] = None

        frames.append(df)

    raw = pd.concat(frames, ignore_index=True)

    # --- Check that all expected columns are present ---
    missing_cols = [c for c in RAW_COLUMNS if c not in raw.columns]
    if missing_cols:
        print(f"  ERROR: The following expected column headers were not found in the CSV(s) for {initials}:")
        for c in missing_cols:
            print(f"    • '{c}'")
        print("  Actual columns found in these files:")
        for c in raw.columns:
            print(f"    • '{c}'")
        print("  Skipping this collector/time block.")
        return None

    # --- Rename columns to standard names ---
    raw = raw.rename(columns=RAW_COLUMNS)

    # --- Convert numeric columns (bad values become NaN rather than crashing) ---
    numeric_cols = ["raw_lat", "raw_lon", "raw_temp", "raw_humidity",
                    "raw_heat_index", "raw_dew_point", "raw_light", "raw_elapsed"]
    for col in numeric_cols:
        raw[col] = pd.to_numeric(raw[col], errors="coerce")

    # --- Convert Time column to 24-hour format ---
    raw["raw_time"] = raw["raw_time"].apply(parse_time_to_24h)

    # --- Round numeric fields to sensible precision ---
    raw["raw_lat"]        = raw["raw_lat"].round(6)
    raw["raw_lon"]        = raw["raw_lon"].round(6)
    raw["raw_temp"]       = raw["raw_temp"].round(2)
    raw["raw_humidity"]   = raw["raw_humidity"].round(2)
    raw["raw_heat_index"] = raw["raw_heat_index"].round(2)
    raw["raw_dew_point"]  = raw["raw_dew_point"].round(2)
    raw["raw_light"]      = raw["raw_light"].round(2)

    # NOTE: Rows with missing latitude/longitude are kept intentionally.
    #       No rows are dropped due to missing GPS coordinates.

    # --- Deduplicate rows with identical timestamps — keep first occurrence ---
    before_dedup = len(raw)
    raw = raw.drop_duplicates(subset=["raw_time"], keep="first").reset_index(drop=True)
    dupes_removed = before_dedup - len(raw)
    if dupes_removed:
        print(f"    Dropped {dupes_removed} duplicate row(s) with identical timestamps.")

    print(f"    Rows after cleaning: {len(raw)}")

    # --- Sort rows chronologically so earliest times appear at the top ---
    raw = raw.sort_values("raw_time", ascending=True).reset_index(drop=True)

    print(f"    Rows after cleaning: {len(raw)}")

    # --- Build the final output table ---
    output = pd.DataFrame({
        "route_id":      ROUTE_ID,
        "route_name":    ROUTE_NAME,
        "route_number":  ROUTE_NUMBER,
        "time_block":    time_block,
        "date":          DATE,
        "time":          raw["raw_time"],
        "latitude":      raw["raw_lat"],
        "longitude":     raw["raw_lon"],
        "temp_probe":    raw["raw_temp"],
        "humidity":      raw["raw_humidity"],
        "dew_point":     raw["raw_dew_point"],
        "heat_index":    raw["raw_heat_index"],
        "ambient_light": raw["raw_light"]
    })

    return output


# =============================================================================
# SECTION 9: DEFINE HELPER — SAVE A DATAFRAME TO A FORMATTED EXCEL FILE
# =============================================================================

def save_to_excel(output_df, initials, time_block):
    """
    Saves a cleaned DataFrame to a formatted Excel file.

    Parameters:
        output_df  : cleaned pandas DataFrame
        initials   : collector initials string (e.g., "BH")
        time_block : "morning" or "afternoon"
    """

    time_label = "AM" if time_block == "morning" else "PM"

    # --- Build output filename ---
    # Format: combined_R{route}_{MMDDYYYY}_{INITIALS}_{AM|PM}.xlsx
    output_filename = os.path.join(
        OUTPUT_DIR_1,
        f"combined_R{ROUTE_NUMBER}_{date_str_for_filename}_{initials}_{time_label}_noGPS.xlsx"
    )

    # --- Build sheet name (Excel sheet names must be 31 characters or fewer) ---
    sheet_name = re.sub(r"[^A-Za-z0-9_]", "_", f"{ROUTE_ID}_{date_str_for_filename}")[:31]

    # --- Write to Excel ---
    output_df.to_excel(output_filename, sheet_name=sheet_name, index=False)

    # --- Apply formatting with openpyxl ---
    wb = load_workbook(output_filename)
    ws = wb[sheet_name]

    header_font   = Font(bold=True)
    header_fill   = PatternFill("solid", fgColor="BDD7EE")
    header_align  = Alignment(horizontal="center")
    header_border = Border(bottom=Side(style="thin"))

    for cell in ws[1]:   # Row 1 = headers
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = header_border

    ws.freeze_panes = "A2"   # Freeze header row so it stays visible when scrolling

    # Auto-size columns to fit content
    for col in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(output_filename)

    print(f"    Saved: {output_filename}")
    print(f"      Sheet : {sheet_name}")
    print(f"      Rows  : {len(output_df)}")
    print(f"      Cols  : {', '.join(output_df.columns)}")


# =============================================================================
# SECTION 10: MAIN LOOP — PROCESS EACH COLLECTOR AND TIME BLOCK
# =============================================================================

if not os.path.isdir(OUTPUT_DIR_1):
    print(f"ERROR: Output folder does not exist: {OUTPUT_DIR_1}")
    print("       Please create this folder before running the script.")
    sys.exit(1)

print("=== Starting processing ===\n")

files_produced = 0

for initials, file_list in sorted(csv_by_initials.items()):

    # Sort each collector's files into morning and afternoon buckets
    am_files = []
    pm_files = []

    for filepath in file_list:
        block = classify_time_block(filepath)
        if block == "morning":
            am_files.append(filepath)
        elif block == "afternoon":
            pm_files.append(filepath)
        else:
            print(f"  WARNING: Could not classify '{os.path.basename(filepath)}' — skipping.")

    # Process and save morning files (if any)
    if am_files:
        print(f"--- {initials} | Morning (AM) ---")
        df_am = process_files(am_files, initials, "morning")
        if df_am is not None:
            save_to_excel(df_am, initials, "morning")
            files_produced += 1
        print()

    # Process and save afternoon files (if any)
    if pm_files:
        print(f"--- {initials} | Afternoon (PM) ---")
        df_pm = process_files(pm_files, initials, "afternoon")
        if df_pm is not None:
            save_to_excel(df_pm, initials, "afternoon")
            files_produced += 1
        print()

print(f"=== Done! {files_produced} Excel file(s) saved to: {OUTPUT_DIR_1} ===")


# =============================================================================
# SECTION 2: FIND ALL EXCEL FILES IN THE INPUT FOLDER
# =============================================================================

EXPECTED_COLUMNS = [
    "route_id", "route_name", "route_number", "time_block",
    "date", "time", "latitude", "longitude", "temp_probe",
    "humidity", "dew_point", "heat_index", "ambient_light"
]

OUTPUT_COLUMNS = [
    "pocketlab", "route_id", "route_name", "route_number", "time_block",
    "date", "time", "latitude", "longitude", "temp_probe",
    "humidity", "dew_point", "heat_index", "ambient_light"
]

if not os.path.isdir(OUTPUT_DIR_1):
    print(f"ERROR: Input folder not found: {OUTPUT_DIR_1}")
    sys.exit(1)

if not os.path.isdir(OUTPUT_DIR):
    print(f"ERROR: Output folder does not exist: {OUTPUT_DIR}")
    print("       Please create this folder before running the script.")
    sys.exit(1)

all_xlsx = sorted([
    f for f in os.listdir(OUTPUT_DIR_1)
    if f.lower().endswith(".xlsx")
])

if not all_xlsx:
    print(f"ERROR: No Excel (.xlsx) files found in folder: {OUTPUT_DIR_1}")
    sys.exit(1)

print(f"=== Found {len(all_xlsx)} Excel file(s) in folder ===")
for f in all_xlsx:
    print(f"  {f}")
print()


# =============================================================================
# SECTION 3: PARSE METADATA FROM EACH FILENAME
# =============================================================================
# Expected filename format:
#   combined_R{route}_{MMDDYYYY}_{PL#}_{AM|PM}_noGPS.xlsx
#
# Examples:
#   combined_R27_06212026_PL1_AM_noGPS.xlsx  → route=27, date=06212026, pocketlab=PL1, block=AM
#   combined_R27_06212026_PL17_PM_noGPS.xlsx → route=27, date=06212026, pocketlab=PL17, block=PM
#   combined_R27_06212026_BH_AM_noGPS.xlsx   → route=27, date=06212026, pocketlab=Unknown, block=AM

FILE_PATTERN = re.compile(
    r"^combined_R(\d+)_(\d{8})_(.+?)_(AM|PM)_noGPS\.xlsx$",
    re.IGNORECASE
)

def parse_filename(filename):
    """
    Parse route number, date string, pocketlab label, and time block from a filename.

    Returns a dict with keys:
        route_number  : int
        date_str      : str  (e.g. "06212026")
        pocketlab     : str  (e.g. "PL1", "PL17", or "Unknown")
        time_block    : str  ("morning" or "afternoon")
        time_label    : str  ("AM" or "PM")

    Returns None if the filename does not match the expected pattern.
    """
    m = FILE_PATTERN.match(filename)
    if not m:
        return None

    route_number = int(m.group(1))
    date_str     = m.group(2)          # e.g. "06212026"
    middle_part  = m.group(3)          # e.g. "PL1", "PL17", "BH"
    time_label   = m.group(4).upper()  # "AM" or "PM"

    # >>> INSERT THE FIX HERE <<<
    if time_label == "AM":
        time_block = "morning"
    else:
        time_block = "afternoon"
    # >>> END OF FIX <<<

    if re.fullmatch(r"PL\d+", middle_part, re.IGNORECASE):
        pocketlab = middle_part.upper()
    else:
        # Build initials from any letter groups in the string
        parts = re.findall(r"[A-Za-z]+", middle_part)
        if parts:
            pocketlab = "".join(parts).upper()  # <<< REMOVED [0] HERE
        else:
            pocketlab = middle_part.upper()

    return {
        "route_number": route_number,
        "date_str":     date_str,
        "pocketlab":    pocketlab,
        "time_block":   time_block,  # This will now successfully find the variable!
        "time_label":   time_label,
    }


parsed_files = []   # List of (filename, metadata_dict)
unparseable  = []   # Filenames that didn't match the pattern

for filename in all_xlsx:
    meta = parse_filename(filename)
    if meta is None:
        unparseable.append(filename)
    else:
        parsed_files.append((filename, meta))

if unparseable:
    print("ERROR: The following file(s) do not match the expected naming format")
    print("       (expected: combined_R{route}_{MMDDYYYY}_{label}_{AM|PM}_noGPS.xlsx):")
    for f in unparseable:
        print(f"  {f}")
    sys.exit(1)


# =============================================================================
# SECTION 4: VALIDATE — ALL FILES MUST SHARE THE SAME ROUTE AND DATE
# =============================================================================

routes_found = set(meta["route_number"] for _, meta in parsed_files)
dates_found  = set(meta["date_str"]     for _, meta in parsed_files)

if len(routes_found) > 1:
    print("ERROR: Files in this folder belong to more than one route.")
    print(f"       Routes found: {sorted(routes_found)}")
    print("       All files in the input folder must share the same route number.")
    sys.exit(1)

if len(dates_found) > 1:
    print("ERROR: Files in this folder belong to more than one date.")
    print(f"       Dates found: {sorted(dates_found)}")
    print("       All files in the input folder must share the same date.")
    sys.exit(1)

ROUTE_NUMBER  = next(iter(routes_found))   # e.g. 27
DATE_STR      = next(iter(dates_found))    # e.g. "06212026"

print(f"=== Validated: all files share route R{ROUTE_NUMBER}, date {DATE_STR} ===")
print()


# =============================================================================
# SECTION 5: VALIDATE — NO DUPLICATE POCKETLAB + TIME BLOCK COMBINATIONS
# =============================================================================

seen_combos = {}   # { (pocketlab, time_block): filename }

for filename, meta in parsed_files:

    # If this file has no PocketLab ID (e.g. CC, KH, etc.),
    # skip the duplicate check.
    if meta["pocketlab"] == "Unknown":
        continue

    key = (meta["pocketlab"], meta["time_block"])

    if key in seen_combos:
        print("ERROR: Duplicate PocketLab + time block combination detected.")
        print(f"       PocketLab : {meta['pocketlab']}")
        print(f"       Time block: {meta['time_block']}")
        print(f"       File 1: {seen_combos[key]}")
        print(f"       File 2: {filename}")
        print("       Each PocketLab must have at most one AM file and one PM file.")
        sys.exit(1)

    seen_combos[key] = filename

print("=== Validated: no duplicate PocketLab + time block combinations ===")
print()


# =============================================================================
# SECTION 6: READ AND COMBINE ALL FILES
# =============================================================================

print("=== Reading files ===")

all_frames = []

for filename, meta in sorted(parsed_files, key=lambda x: (x[1]["pocketlab"], x[1]["time_label"])):
    filepath = os.path.join(OUTPUT_DIR_1, filename)
    print(f"  Reading: {filename}")
    print(f"    PocketLab  : {meta['pocketlab']}")
    print(f"    Time block : {meta['time_label']}")

    try:
        df = pd.read_excel(filepath)
    except Exception as e:
        print(f"  ERROR: Could not read file '{filename}': {e}")
        sys.exit(1)

    # --- Check that all expected columns are present ---
    missing = [c for c in EXPECTED_COLUMNS if c not in df.columns]
    if missing:
        print(f"  ERROR: The following expected columns are missing from '{filename}':")
        for c in missing:
            print(f"    • '{c}'")
        print("  Columns actually found in this file:")
        for c in df.columns:
            print(f"    • '{c}'")
        sys.exit(1)

    # --- Coerce numeric columns to proper number types ---
    numeric_cols = ["route_number", "latitude", "longitude", "temp_probe",
                    "humidity", "dew_point", "heat_index", "ambient_light"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # --- Ensure text columns are stored as strings, not mixed types ---
    text_cols = ["route_id", "route_name", "time_block", "date", "time"]
    for col in text_cols:
        df[col] = df[col].astype(str)

    # --- Add the pocketlab column at the front ---
    df.insert(0, "pocketlab", meta["pocketlab"])

    # --- Enforce output column order ---
    df = df[OUTPUT_COLUMNS]

    print(f"    Rows read  : {len(df)}")
    all_frames.append(df)
    print()

combined = pd.concat(all_frames, ignore_index=True)

print(f"=== Combined total rows: {len(combined)} ===")
print()


# =============================================================================
# SECTION 7: SAVE COMBINED OUTPUT TO EXCEL
# =============================================================================

output_filename_base = f"full_R{ROUTE_NUMBER}_{DATE_STR}"
output_filepath      = os.path.join(OUTPUT_DIR, output_filename_base + ".xlsx")

# Excel sheet names must be 31 characters or fewer
sheet_name = output_filename_base[:31]

print(f"=== Saving output ===")
print(f"  File  : {output_filepath}")
print(f"  Sheet : {sheet_name}")

try:
    combined.to_excel(output_filepath, sheet_name=sheet_name, index=False)
except Exception as e:
    print(f"ERROR: Could not save output file: {e}")
    sys.exit(1)

print(f"  Rows  : {len(combined)}")
print(f"  Cols  : {', '.join(OUTPUT_COLUMNS)}")
print()
print(f"=== Done! Combined file saved to: {output_filepath} ===")